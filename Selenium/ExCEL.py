import openpyxl
import time
#For Excle the steps are->
#1. load the excel
book = openpyxl.load_workbook("C:\\Users\\LJA095\\OneDrive - Maersk Group\\Documents\\Neonnav L2 To L3\\L2 to L3 Auto rerouting.xlsx")
time.sleep(5)
sheet = book.active
#2. active sheet
#sheet = workbook["Sheet1"]
time.sleep(1)
#3. getting data
data = sheet.cell(row = 2, column = 3)
print(data.value)
csr = data.value
total_rows = sheet.max_row
#5. write in excle

sheet.cell(2,26).value="Lakshay Jain"

book.save("C:\\Users\\LJA095\\OneDrive - Maersk Group\\Documents\\Neonnav L2 To L3\\L2 to L3 Auto rerouting.xlsx")
