
from openpyxl.workbook import workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import time
from selenium.webdriver.common.keys import Keys

# chrome browser
driver = webdriver.Chrome()
driver2 = webdriver.Chrome()
driver.maximize_window()
time.sleep(2)
#double browser instance
browser = driver.get("https://e2open.my.site.com/support/s/login/")
#browser = driver2.get("https://www.google.com/search?q=power+automate&sca_esv=0c2c4a4daad34013&source=hp&ei=URORZtuGGaTe2roPrcSc0AM&iflsig=AL9hbdgAAAAAZpEhYc7KqM-DeBfsLgHZ8zHNfLgpg5Qy&ved=0ahUKEwib85PosqGHAxUkr1YBHS0iBzoQ4dUDCA0&uact=5&oq=power+automate&gs_lp=Egdnd3Mtd2l6Ig5wb3dlciBhdXRvbWF0ZTILEAAYgAQYsQMYgwEyCxAAGIAEGLEDGIMBMgUQABiABDIFEAAYgAQyCxAAGIAEGLEDGIMBMggQABiABBixAzIFEAAYgAQyCxAAGIAEGLEDGIMBMgUQABiABDIFEAAYgARI5C9QDFivLXABeACQAQGYAeIGoAHAG6oBDTAuMS4xLjAuMS4zLjG4AQPIAQD4AQGYAgegAoMVqAIAwgINEAAYgAQYsQMYgwEYCsICBxAAGIAEGArCAgoQABiABBixAxgKmAMAkgcLMS4xLjEuMC4yLjKgB-0q&sclient=gws-wiz")
time.sleep(2)
#login page
driver.find_element(By.XPATH,"//input[@id='48:2;a']").send_keys("abhinay.waghmare@maersk.com")
time.sleep(5)
driver.find_element(By.XPATH,"//input[@id='61:2;a']").send_keys("Aishabhi29039397@")
time.sleep(5)
driver.find_element(By.XPATH,"//span[@class=' label bBody']").click()
time.sleep(20)
#//////////////////////////////////////////////////////////////////////////////////////////////////

#For Excle the steps are->
#1. load the excel
book = openpyxl.load_workbook("C:\\Users\\LJA095\\OneDrive - Maersk Group\\Documents\\Neonnav L2 To L3\\L2 to L3 Auto rerouting.xlsx")
time.sleep(5)
#Sample1 = openpyxl.load_workbook("C:\\Users\\LJA095\\.vscode\\Selenium\\Sample.xlsx")
time.sleep(5)
#sheet_sample = Sample1.active
#2. active sheet
#sheet = workbook["Sheet1"]
sheet = book.active
time.sleep(1)
arr = []
#3. getting data
data = sheet.cell(row = 2, column = 3)
print(data.value)
csr = data.value
#4. total no of rows
total_rows = sheet.max_row
time.sleep(5)
#//////////////////////////////////////////////////////////////////////
# website actions
driver.find_element(By.XPATH,"//input[@id='132:0-input']").send_keys(csr)
time.sleep(5)
driver.find_element(By.XPATH,"//button[@class='slds-button slds-button_icon search-inputSearchButton search-inputSearchButton--right slds-button_icon-border']").click()
time.sleep(15)
#read from website
text1 =driver.find_element(By.XPATH,"//tbody//tr//td[3]").text
#5. write in excle
sheet.cell(2,5).value=text1
book.save("C:\\Users\\LJA095\\OneDrive - Maersk Group\\Documents\\Neonnav L2 To L3\\L2 to L3 Auto rerouting.xlsx")
print(text1)
arr.append(text1)
count = 0
#total_rows + 1
for i in range(3, 10 , 1):
    driver = webdriver.Chrome()
    driver.maximize_window()
    time.sleep(2)

    browser = driver.get("https://e2open.my.site.com/support/s/")
    time.sleep(5)
    
    driver.find_element(By.XPATH,"//input[@id='48:2;a']").send_keys("abhinay.waghmare@maersk.com")
    time.sleep(5)

    driver.find_element(By.XPATH,"//input[@id='61:2;a']").send_keys("Aishabhi29039397@")
    time.sleep(5)

    driver.find_element(By.XPATH,"//span[@class=' label bBody']").click()
    time.sleep(20)

    
    data = sheet.cell(row = i, column = 3)
    print(data.value)
    csr = data.value
    time.sleep(5)

    driver.find_element(By.XPATH,"//input[@id='132:0-input']").send_keys(csr)
    time.sleep(5)
    driver.find_element(By.XPATH,"//button[@class='slds-button slds-button_icon search-inputSearchButton search-inputSearchButton--right slds-button_icon-border']").click()
    time.sleep(15)
    
    text1 =driver.find_element(By.XPATH,"//tbody//tr//td[3]").text
    print(text1)
    sheet.cell(i,5).value=text1
    book.save("C:\\Users\\LJA095\\OneDrive - Maersk Group\\Documents\\Neonnav L2 To L3\\L2 to L3 Auto rerouting.xlsx")
    print(text1)
    count = count + 1
    if text1=="":
        arr.append("NA")
    else:
        arr.append(text1)
    time.sleep(1)
    print(arr)
    print(count)

print(arr)


