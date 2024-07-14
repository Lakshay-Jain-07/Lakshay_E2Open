from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl
import time
from selenium.webdriver.common.keys import Keys
from openpyxl.workbook import workbook
import os
import time
import shutil


def get_latest_downloaded_file(download_folder):
    # Get the latest downloaded file from the download folder
    files = os.listdir(download_folder)
    paths = [os.path.join(download_folder, basename) for basename in files if basename.endswith('.crdownload') is False]
    latest_file = max(paths, key=os.path.getctime)
    return latest_file

def rename_and_move_file(source, destination_folder, new_name, folder_name):
    # Create a new folder in the destination folder
    new_folder_path = os.path.join(destination_folder, 'new_folder')
    if not os.path.exists(new_folder_path):
        os.makedirs(new_folder_path)
        print(f"Created folder: {new_folder_path}")
    
    # Create the full destination path
    destination_path = os.path.join(new_folder_path, new_name)
    
    # Rename and move the file
    shutil.move(source, destination_path)
    print(f"File moved to {destination_path}")


#For Excle the steps are->
#1. load the excel
book = openpyxl.load_workbook("C:\\Users\\LJA095\\Maersk Group\\Test BOT group - Bicmapping\\EDI Resource Attachment .xlsx")
time.sleep(5)
#Sample1 = openpyxl.load_workbook("C:\\Users\\LJA095\\.vscode\\Selenium\\Sample.xlsx")
time.sleep(5)
#sheet_sample = Sample1.active
#2. active sheet
#sheet = workbook["Sheet1"]
sheet = book.active
time.sleep(1)
FFR = sheet.max_row
for i in range(2, FFR+1, 1):
    Track_ID = sheet.cell(row = i, column = 2)
    MSK_ID =Track_ID.value
    print(MSK_ID)
    Incident = sheet.cell(row=i,column=1)
#Browser launch
    driver = webdriver.Chrome()
    driver.maximize_window()
    time.sleep(2)
    browser = driver.get("https://portal.maersk.seeburger-ipaas.cloud/suite/app/?organization=MSK#!BT1~bismtMSK-defaultMTmenu.1-defaultMTmenu.1.s.s.defaultworkflow/selectedTabIndex=0/seed=/defaultMTmenu.1.s.defaultworkflow/selectedTabIndex=0,")
    time.sleep(2)
#page 1 actions
    driver.find_element(By.XPATH,"//span[normalize-space()='Clear Values']").click()
    time.sleep(5)
    driver.find_element(By.XPATH,"(//input[@id='26394237EFE788242A5C0C029B98F802562522F0-searchText1']").send_keys(MSK_ID)
    time.sleep(20)
    driver.find_element(By.XPATH,"(//div[@role='button'])[14]").send_keys("Is equal to" + Keys.RETURN)
    #driver.find_element(By.XPATH,"(//div[@role='button'])[14]").send_keys(Keys.RETURN)
    time.sleep(2)
    driver.find_element(By.XPATH,"//span[contains(text(),'Search')]").click()
    time.sleep(5)
#page 2 actions
    driver.find_element(By.XPATH,"//label[@for='gwt-uid-70']").click()
    time.sleep(1)
    driver.find_element(By.CSS_SELECTOR,"body > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(9) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(3) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(3) > div:nth-child(1) > span:nth-child(2)").click()
    time.sleep(1)
    driver.find_element(By.XPATH,"//span[contains(text(),'Source View')]").click()
    time.sleep(2)
    driver.find_element(By.XPATH,"//span[@class='v-button-caption'][normalize-space()='Download']").click()
    time.sleep(2)
# Downloading attachment
    # Path to your download folder
    download_folder = "C:\\Users\\LJA095\\Downloads"
    # Desired new name for the file
    new_name = MSK_ID
    # Path to the destination folder
    destination_folder = "C:\\Users\\LJA095\OneDrive - Maersk Group\\Documents\\Attachments"
    time.sleep(10)
    latest_file = get_latest_downloaded_file(download_folder)
    # latest downloaded file
    print(f"Latest downloaded file: {latest_file}")
# Rename and move the file
    rename_and_move_file(latest_file, destination_folder, new_name, Incident)






